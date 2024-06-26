import os
import pandas as pd
import openpyxl

# Define the base path where the folders are located
base_path = 'C:\\Users\\vbaAp\\Desktop\\Indeksacja'  # Replace with the base path of your folders

# Load the CSV file
csv_file_path = 'C:\\Users\\vbaAp\\Desktop\\Indeksacja\\checkContract.csv'  # Replace with the path to your CSV file
df = pd.read_csv(csv_file_path, delimiter=';')

# Print the columns to diagnose the issue
print("Columns in CSV file:", df.columns)

# # Strip any leading/trailing spaces from column names
# df.columns = df.columns.str.strip()

# # Check again after stripping spaces
# print("Columns after stripping spaces:", df.columns)

# Function to find the row with the value 'Coverage' in column A
def find_coverage_row(sheet):
    for row in sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value == 'Coverage':
                return cell.row
    return None

# Loop through each row in the CSV file
for index, row in df.iterrows():
    vin = row['VIN']
    client_version = row['clientVersionFile']
    dealer_version = row['dealerVersionFile']
    
    client_file_path = os.path.join(base_path, client_version, str(vin), '2024', f'{vin}.xlsx')
    dealer_file_path = os.path.join(base_path, dealer_version, str(vin), '2024', f'{vin}.xlsx')

    # print("client path: ", client_file_path)
    # print("dealer path: ", client_file_path)

    if os.path.exists(client_file_path) and os.path.exists(dealer_file_path):
        # Load the Excel files
        client_wb = openpyxl.load_workbook(client_file_path)
        dealer_wb = openpyxl.load_workbook(dealer_file_path)
        
        client_sheet = client_wb.active
        dealer_sheet = dealer_wb.active
        
        # Find the row with 'Coverage' in column A
        client_coverage_row = find_coverage_row(client_sheet)
        dealer_coverage_row = find_coverage_row(dealer_sheet)
        
        if client_coverage_row and dealer_coverage_row:
            # Get the values in column D for the found row
            client_value = client_sheet.cell(row=client_coverage_row, column=4).value
            dealer_value = dealer_sheet.cell(row=dealer_coverage_row, column=4).value
            
            # Compare the values
            if client_value == dealer_value:
                print(f'For VIN {vin}, the values in row {client_coverage_row} and column D are the same: {client_value}')
            else:
                print(f'For VIN {vin}, the values in row {client_coverage_row} and column D are different. Client: {client_value}, Dealer: {dealer_value}')
        else:
            print(f'Coverage row not found in one or both files for VIN {vin}.')
    else:
        print(f'One or both files for VIN {vin} do not exist.')

print('Comparison complete!')
